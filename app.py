from __future__ import annotations

import argparse
import html
import json
import mimetypes
import re
import sys
import traceback
import uuid
import zipfile
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from email.parser import BytesParser
from email.policy import default
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote, urlparse

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


APP_ROOT = Path(__file__).resolve().parent
UPLOAD_DIR = APP_ROOT / "uploads"
OUTPUT_DIR = APP_ROOT / "outputs"

SOURCE_HEADERS = {
    "order_no": "订单号",
    "customer": "客户名称",
    "delivery_date": "发货日期",
    "product": "商品名称",
    "description": "商品描述",
    "quantity": "实际数量",
    "unit": "发货单位",
    "price": "发货单价",
    "amount": "实际金额",
}

TABLE_HEADERS = ["发货时间", "商品名称", "商品描述", "实际数量", "单位", "单价（元）", "实际金额（元）", "日小计（元）"]


@dataclass
class OrderRow:
    row_no: int
    order_no: str
    customer: str
    delivery_date: date
    product: str
    description: str | None
    quantity: Any
    unit: str | None
    price: Any
    amount: Decimal
    meal_type: str


@dataclass
class GeneratedSheet:
    name: str
    meal_type: str
    year: int
    month: int
    rows: int
    days: int
    total: Decimal


@dataclass
class GeneratedWorkbook:
    output_name: str
    school: str
    template_name: str
    sheets: list[GeneratedSheet]
    warnings: list[str]


class WorkbookProcessError(Exception):
    pass


def html_page(title: str, body: str) -> bytes:
    page = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --ink: #1f2933;
      --muted: #667085;
      --line: #d7dce2;
      --accent: #0f766e;
      --accent-2: #164e63;
      --danger: #b42318;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--ink);
      font: 14px/1.5 -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft YaHei", sans-serif;
    }}
    .shell {{
      max-width: 1080px;
      margin: 0 auto;
      padding: 28px 20px 48px;
    }}
    header {{
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 20px;
      margin-bottom: 18px;
    }}
    h1 {{
      margin: 0;
      font-size: 24px;
      line-height: 1.25;
      font-weight: 700;
    }}
    .sub {{
      margin: 6px 0 0;
      color: var(--muted);
      max-width: 740px;
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 20px;
      box-shadow: 0 1px 2px rgba(16, 24, 40, 0.04);
    }}
    form {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 16px;
    }}
    label {{
      display: block;
      font-weight: 650;
      margin-bottom: 7px;
    }}
    input[type="file"], input[type="text"] {{
      width: 100%;
      min-height: 40px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      padding: 8px 10px;
      color: var(--ink);
    }}
    input[type="file"] {{ padding: 7px; }}
    .full {{ grid-column: 1 / -1; }}
    .hint {{
      margin-top: 6px;
      color: var(--muted);
      font-size: 12px;
    }}
    .checks {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px 18px;
      padding: 12px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fbfcfd;
    }}
    .checks label {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      margin: 0;
      font-weight: 500;
    }}
    .actions {{
      display: flex;
      align-items: center;
      gap: 12px;
      margin-top: 4px;
    }}
    button, .button {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 40px;
      border: 1px solid var(--accent);
      border-radius: 6px;
      padding: 0 15px;
      background: var(--accent);
      color: #fff;
      text-decoration: none;
      font-weight: 650;
      cursor: pointer;
    }}
    .secondary {{
      border-color: var(--line);
      background: #fff;
      color: var(--ink);
    }}
    .danger {{
      border-color: #b42318;
      background: #b42318;
      color: #fff;
    }}
    .top-grid {{
      display: grid;
      grid-template-columns: 1fr;
      gap: 16px;
    }}
    .history-head {{
      display: flex;
      justify-content: space-between;
      align-items: baseline;
      gap: 12px;
      margin-bottom: 8px;
    }}
    .history-head h2 {{
      margin: 0;
      font-size: 18px;
      line-height: 1.3;
    }}
    .badge {{
      display: inline-block;
      border: 1px solid #b7c7d8;
      border-radius: 999px;
      padding: 1px 8px;
      color: #164e63;
      font-size: 12px;
      font-weight: 700;
      white-space: nowrap;
    }}
    .file-actions {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
    }}
    .file-actions form {{
      display: inline;
    }}
    .file-actions button, .file-actions .button {{
      min-height: 32px;
      padding: 0 10px;
      font-size: 13px;
    }}
    .empty {{
      color: var(--muted);
      margin: 6px 0 0;
    }}
    .message {{
      padding: 12px 14px;
      border-radius: 6px;
      border: 1px solid var(--line);
      background: #fbfcfd;
      margin-bottom: 16px;
    }}
    .error {{
      border-color: #f1b8b3;
      background: #fff6f5;
      color: var(--danger);
      white-space: pre-wrap;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      margin-top: 12px;
      font-size: 13px;
    }}
    th, td {{
      border-bottom: 1px solid var(--line);
      text-align: left;
      padding: 8px 10px;
      vertical-align: top;
    }}
    th {{
      background: #f2f5f7;
      color: #344054;
      font-weight: 700;
    }}
    .total {{
      font-weight: 700;
      color: var(--accent-2);
    }}
    code {{
      font-family: ui-monospace, SFMono-Regular, Consolas, "Liberation Mono", monospace;
      font-size: 12px;
      background: #eef2f6;
      padding: 2px 5px;
      border-radius: 4px;
    }}
    @media (max-width: 760px) {{
      .shell {{ padding: 18px 12px 32px; }}
      header {{ display: block; }}
      form {{ grid-template-columns: 1fr; }}
      .actions {{ flex-direction: column; align-items: stretch; }}
      button, .button {{ width: 100%; }}
      .file-actions {{ flex-direction: column; }}
    }}
  </style>
</head>
<body>
  <main class="shell">
    <header>
      <div>
        <h1>配送月清单导入工具</h1>
        <p class="sub">上传订单导出表和学校月清单模板，自动按学校、日期和餐别生成新的月份 sheet，保留模板样式。</p>
      </div>
    </header>
    {body}
  </main>
</body>
</html>"""
    return page.encode("utf-8")


def format_file_size(size: int) -> str:
    if size < 1024:
        return f"{size} B"
    if size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    return f"{size / (1024 * 1024):.1f} MB"


def safe_filename(value: str, fallback: str = "output") -> str:
    text = clean_text(value) or fallback
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", text)
    text = re.sub(r"\s+", "_", text).strip("._ ")
    return (text or fallback)[:80]


def list_saved_outputs() -> list[dict[str, Any]]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    files = []
    for path in OUTPUT_DIR.glob("*.xlsx"):
        if not path.is_file():
            continue
        stat = path.stat()
        files.append(
            {
                "name": path.name,
                "timestamp": stat.st_mtime,
                "mtime": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                "size": format_file_size(stat.st_size),
            }
        )
    return sorted(files, key=lambda item: item["timestamp"], reverse=True)


def render_saved_outputs() -> str:
    saved = list_saved_outputs()
    if not saved:
        return """
<section class="panel">
  <div class="history-head">
    <h2>已保存的目标清单</h2>
  </div>
  <p class="empty">还没有生成过目标清单。</p>
</section>"""

    rows = []
    for index, item in enumerate(saved):
        name = item["name"]
        latest = '<span class="badge">最新</span>' if index == 0 else ""
        rows.append(
            "<tr>"
            f"<td>{latest}</td>"
            f"<td>{html.escape(name)}</td>"
            f"<td>{html.escape(item['mtime'])}</td>"
            f"<td>{html.escape(item['size'])}</td>"
            '<td><div class="file-actions">'
            f'<a class="button secondary" href="/download/{quote(name)}">下载</a>'
            f'<form action="/delete/{quote(name)}" method="post" onsubmit="return confirm(\'确定删除这个目标清单吗？删除后不能从工具中恢复。\');">'
            '<button class="danger" type="submit">删除</button>'
            "</form>"
            "</div></td>"
            "</tr>"
        )
    return f"""
<section class="panel">
  <div class="history-head">
    <h2>已保存的目标清单</h2>
    <span class="hint">除非删除，否则会一直保存在本机 outputs 文件夹。</span>
  </div>
  <table>
    <thead><tr><th></th><th>文件名</th><th>更新时间</th><th>大小</th><th>操作</th></tr></thead>
    <tbody>{''.join(rows)}</tbody>
  </table>
</section>"""


def render_form(message: str = "") -> bytes:
    escaped = f'<div class="message">{html.escape(message)}</div>' if message else ""
    body = f"""{escaped}
<div class="top-grid">
{render_saved_outputs()}
<section class="panel">
  <form action="/process" method="post" enctype="multipart/form-data">
    <div>
      <label for="source_file">订单导出文件</label>
      <input id="source_file" name="source_file" type="file" accept=".xlsx" multiple required>
      <div class="hint">可一次选择多个订单导出文件；格式需包含：订单号、客户名称、发货日期、商品名称、实际数量等列。</div>
    </div>
    <div>
      <label for="template_file">目标学校月清单模板</label>
      <input id="template_file" name="template_file" type="file" accept=".xlsx" multiple required>
      <div class="hint">可一次选择多个学校模板；每个模板会从“配送学校”单元格自动识别学校名并单独生成结果。</div>
    </div>
    <div class="full">
      <label for="school_name">学校名称（可选）</label>
      <input id="school_name" name="school_name" type="text" placeholder="不填则自动读取模板里的配送学校">
      <div class="hint">仅单个目标模板时使用；如果源文件学校名和模板不完全一致，可以在这里手动填写源文件中的学校名。</div>
    </div>
    <div class="full checks">
      <label><input type="checkbox" name="include_regular" checked> 普通营养餐</label>
      <label><input type="checkbox" name="include_evening" checked> 早晚餐</label>
      <label><input type="checkbox" name="replace_existing"> 同名月份 sheet 存在时覆盖</label>
    </div>
    <div class="full actions">
      <button type="submit">生成清单</button>
      <span class="hint">批量模式会按目标学校分别生成结果；默认不覆盖已有 sheet，会生成“导入”后缀。</span>
    </div>
  </form>
</section>
</div>"""
    return html_page("配送月清单导入工具", body)


def create_batch_zip(output_names: list[str], zip_name: str) -> Path:
    zip_path = OUTPUT_DIR / zip_name
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for output_name in output_names:
            path = get_output_path(output_name, allowed_suffixes=(".xlsx",))
            if path.exists():
                archive.write(path, arcname=path.name)
    return zip_path


def render_batch_result(generated: list[GeneratedWorkbook], errors: list[str], zip_name: str | None) -> bytes:
    rows = []
    downloadable = []
    for item in generated:
        sheet_summary = "；".join(
            f"{sheet.name} {sheet.rows}行 {format_decimal(sheet.total)}元"
            for sheet in item.sheets
        )
        total_rows = sum(sheet.rows for sheet in item.sheets)
        total_amount = sum((sheet.total for sheet in item.sheets), Decimal("0"))
        downloadable.append(item.output_name)
        rows.append(
            "<tr>"
            f"<td>{html.escape(item.school)}</td>"
            f"<td>{html.escape(item.output_name)}</td>"
            f"<td>{total_rows}</td>"
            f"<td class=\"total\">{format_decimal(total_amount)}</td>"
            f"<td>{html.escape(sheet_summary)}</td>"
            '<td><div class="file-actions">'
            f'<a class="button secondary" href="/download/{quote(item.output_name)}">下载</a>'
            "</div></td>"
            "</tr>"
        )

    warning_items = []
    for item in generated:
        for warning in item.warnings:
            warning_items.append(f"{item.school}：{warning}")
    warning_items.extend(errors)
    warning_html = ""
    if warning_items:
        warning_html = "<div class=\"message\"><strong>提示</strong><ul>" + "".join(
            f"<li>{html.escape(message)}</li>" for message in warning_items
        ) + "</ul></div>"

    zip_link = f'<a class="button" href="/download/{quote(zip_name)}">下载全部 ZIP</a>' if zip_name else ""
    folder_button = ""
    if downloadable:
        files_json = html.escape(json.dumps(downloadable, ensure_ascii=False), quote=True)
        folder_button = (
            f'<button class="button secondary" type="button" onclick="saveBatchToFolder({files_json})">'
            "选择文件夹保存</button>"
        )

    body = f"""
<section class="panel">
  <div class="message">已生成 <strong>{len(generated)}</strong> 个目标学校清单。</div>
  {warning_html}
  <div class="actions">
    {zip_link}
    {folder_button}
    <a class="button secondary" href="/">继续处理</a>
  </div>
  <table>
    <thead><tr><th>学校</th><th>结果文件</th><th>明细行</th><th>合计金额</th><th>sheet 明细</th><th>操作</th></tr></thead>
    <tbody>{''.join(rows)}</tbody>
  </table>
</section>
<script>
async function saveBatchToFolder(files) {{
  if (!window.showDirectoryPicker) {{
    alert("当前浏览器或当前访问方式不支持直接选择文件夹。请使用“下载全部 ZIP”。");
    return;
  }}
  const dir = await window.showDirectoryPicker();
  for (const name of files) {{
    const response = await fetch("/download/" + encodeURIComponent(name));
    if (!response.ok) throw new Error("下载失败：" + name);
    const blob = await response.blob();
    const handle = await dir.getFileHandle(name, {{ create: true }});
    const writable = await handle.createWritable();
    await writable.write(blob);
    await writable.close();
  }}
  alert("保存完成。");
}}
</script>"""
    return html_page("生成完成", body)


def render_result(output_name: str, school: str, sheets: list[GeneratedSheet], warnings: list[str]) -> bytes:
    rows = []
    for sheet in sheets:
        meal = "早晚餐" if sheet.meal_type == "evening" else "普通营养餐"
        rows.append(
            "<tr>"
            f"<td>{html.escape(sheet.name)}</td>"
            f"<td>{meal}</td>"
            f"<td>{sheet.year}年{sheet.month}月</td>"
            f"<td>{sheet.days}</td>"
            f"<td>{sheet.rows}</td>"
            f"<td class=\"total\">{format_decimal(sheet.total)}</td>"
            "</tr>"
        )
    warning_html = ""
    if warnings:
        warning_items = "".join(f"<li>{html.escape(w)}</li>" for w in warnings)
        warning_html = f'<div class="message"><strong>提示</strong><ul>{warning_items}</ul></div>'
    table = "".join(rows)
    link = f"/download/{quote(output_name)}"
    body = f"""
<section class="panel">
  <div class="message">已生成：<strong>{html.escape(output_name)}</strong><br>配送学校：{html.escape(school)}</div>
  {warning_html}
  <div class="actions">
    <a class="button" href="{link}">下载结果 Excel</a>
    <a class="button secondary" href="/">继续处理</a>
  </div>
  <table>
    <thead><tr><th>sheet</th><th>餐别</th><th>月份</th><th>日期数</th><th>明细行</th><th>合计金额</th></tr></thead>
    <tbody>{table}</tbody>
  </table>
</section>"""
    return html_page("生成完成", body)


def render_error(error: str) -> bytes:
    body = f"""
<section class="panel">
  <div class="message error">{html.escape(error)}</div>
  <div class="actions"><a class="button secondary" href="/">返回重新上传</a></div>
</section>"""
    return html_page("处理失败", body)


def get_output_path(name: str, allowed_suffixes: tuple[str, ...] = (".xlsx",)) -> Path:
    safe_name = Path(name).name
    path = (OUTPUT_DIR / safe_name).resolve()
    output_root = OUTPUT_DIR.resolve()
    if path.parent != output_root or path.suffix.lower() not in allowed_suffixes:
        allowed = "、".join(allowed_suffixes)
        raise WorkbookProcessError(f"只能操作已生成的 {allowed} 文件。")
    return path


def get_saved_output_path(name: str) -> Path:
    return get_output_path(name, allowed_suffixes=(".xlsx",))


def parse_multipart(headers: Any, body: bytes) -> tuple[dict[str, str], dict[str, list[tuple[str, bytes]]]]:
    content_type = headers.get("Content-Type", "")
    raw = f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8") + body
    message = BytesParser(policy=default).parsebytes(raw)
    if not message.is_multipart():
        raise WorkbookProcessError("上传请求格式不正确。")

    fields: dict[str, str] = {}
    files: dict[str, list[tuple[str, bytes]]] = {}
    for part in message.iter_parts():
        disposition = part.get("Content-Disposition", "")
        if "form-data" not in disposition:
            continue
        name = part.get_param("name", header="Content-Disposition")
        filename = part.get_filename()
        payload = part.get_payload(decode=True) or b""
        if not name:
            continue
        if filename:
            files.setdefault(name, []).append((Path(filename).name, payload))
        else:
            fields[name] = payload.decode(part.get_content_charset() or "utf-8", errors="replace")
    return fields, files


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    match = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", text)
    if match:
        y, m, d = map(int, match.groups())
        return date(y, m, d)
    return None


def to_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def decimal_to_cell(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def format_decimal(value: Decimal) -> str:
    normalized = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{normalized:,.2f}"


def zh_date(d: date) -> str:
    return f"{d.year}年{d.month}月{d.day}日"


def safe_sheet_name(base: str, workbook: Any, replace_existing: bool) -> str:
    invalid = r"[\[\]\:\*\?\/\\]"
    base = re.sub(invalid, "_", base)[:31]
    existing_names = {name.strip() for name in workbook.sheetnames}
    if replace_existing or base.strip() not in existing_names:
        return base

    suffix = " 导入"
    candidate = (base[: 31 - len(suffix)] + suffix)[:31]
    if candidate not in workbook.sheetnames:
        return candidate
    for i in range(2, 100):
        suffix = f" 导入{i}"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
        if candidate not in workbook.sheetnames:
            return candidate
    raise WorkbookProcessError(f"无法为 {base} 生成唯一的 sheet 名。")


def find_sheet_by_title(workbook: Any, title: str) -> Worksheet | None:
    wanted = title.strip()
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip() == wanted:
            return workbook[sheet_name]
    return None


def find_source_table(workbook: Any) -> tuple[Worksheet, int, dict[str, int]]:
    required = set(SOURCE_HEADERS.values())
    for sheet in workbook.worksheets:
        max_scan = min(sheet.max_row or 1, 20)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
            headers = [clean_text(cell) for cell in row]
            if required.issubset(set(headers)):
                index = {name: headers.index(name) for name in required}
                return sheet, row_idx, index
    raise WorkbookProcessError("订单导出文件中没有找到预期表头。")


def read_orders(path: Path, school_name: str, include_regular: bool, include_evening: bool) -> list[OrderRow]:
    source_workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet, header_row, index = find_source_table(source_workbook)
        school_key = normalize_text(school_name)
        evening_key = normalize_text(f"{school_name}-早晚餐")
        orders: list[OrderRow] = []

        for row_no, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            customer = clean_text(row[index[SOURCE_HEADERS["customer"]]])
            customer_key = normalize_text(customer)
            meal_type = ""
            if include_regular and customer_key == school_key:
                meal_type = "regular"
            elif include_evening and customer_key == evening_key:
                meal_type = "evening"
            else:
                continue

            delivery_date = parse_date(row[index[SOURCE_HEADERS["delivery_date"]]])
            product = clean_text(row[index[SOURCE_HEADERS["product"]]])
            if not delivery_date or not product:
                continue

            quantity = row[index[SOURCE_HEADERS["quantity"]]]
            price = row[index[SOURCE_HEADERS["price"]]]
            amount = to_decimal(row[index[SOURCE_HEADERS["amount"]]])
            if amount == 0 and quantity not in (None, "") and price not in (None, ""):
                amount = to_decimal(quantity) * to_decimal(price)

            desc = clean_text(row[index[SOURCE_HEADERS["description"]]]) or None
            orders.append(
                OrderRow(
                    row_no=row_no,
                    order_no=clean_text(row[index[SOURCE_HEADERS["order_no"]]]),
                    customer=customer,
                    delivery_date=delivery_date,
                    product=product,
                    description=desc,
                    quantity=quantity,
                    unit=clean_text(row[index[SOURCE_HEADERS["unit"]]]) or None,
                    price=price,
                    amount=amount,
                    meal_type=meal_type,
                )
            )

        return sorted(orders, key=lambda item: (item.meal_type, item.delivery_date, item.order_no, item.row_no))
    finally:
        source_workbook.close()


def read_orders_from_sources(source_paths: list[Path], school_name: str, include_regular: bool, include_evening: bool) -> list[OrderRow]:
    orders: list[OrderRow] = []
    for source_path in source_paths:
        orders.extend(read_orders(source_path, school_name, include_regular, include_evening))
    return sorted(orders, key=lambda item: (item.meal_type, item.delivery_date, item.order_no, item.row_no))


def extract_school_name(workbook: Any) -> str:
    pattern = re.compile(r"配送学校\s*[:：]\s*(.+)")
    for sheet in workbook.worksheets:
        max_row = min(sheet.max_row or 1, 8)
        max_col = min(sheet.max_column or 1, 8)
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                value = clean_text(cell.value)
                match = pattern.search(value)
                if match:
                    return match.group(1).strip()
    raise WorkbookProcessError("目标模板中没有找到“配送学校：...”信息，请在页面里手动填写学校名称。")


def parse_sheet_month(name: str) -> tuple[int, int] | None:
    match = re.search(r"(\d{4})年\s*(\d{1,2})月", name)
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def sheet_meal_type(sheet: Worksheet) -> str:
    name = sheet.title
    title = clean_text(sheet.cell(1, 1).value)
    if "早晚餐" in name or "早晚餐" in title:
        return "evening"
    return "regular"


def choose_template_sheet(workbook: Any, meal_type: str, year: int, month: int) -> Worksheet:
    exact_name = base_sheet_name(meal_type, year, month)
    exact_sheet = find_sheet_by_title(workbook, exact_name)
    if exact_sheet is not None:
        return exact_sheet

    candidates: list[tuple[int, int, int, Worksheet]] = []
    for idx, sheet in enumerate(workbook.worksheets):
        if sheet_meal_type(sheet) != meal_type:
            continue
        parsed = parse_sheet_month(sheet.title)
        if parsed:
            candidates.append((parsed[0], parsed[1], idx, sheet))
    if candidates:
        candidates.sort()
        return candidates[-1][3]

    readable = "早晚餐" if meal_type == "evening" else "普通营养餐"
    raise WorkbookProcessError(f"目标模板中没有找到可复制的{readable} sheet。")


def base_sheet_name(meal_type: str, year: int, month: int) -> str:
    if meal_type == "evening":
        return f"{year}年{month}月 早晚餐"
    return f"{year}年{month}月"


def find_table_layout(sheet: Worksheet) -> tuple[int, int, int]:
    header_row = None
    for row_idx in range(1, min(sheet.max_row or 1, 20) + 1):
        values = [clean_text(sheet.cell(row_idx, col).value) for col in range(1, 9)]
        if values[: len(TABLE_HEADERS)] == TABLE_HEADERS:
            header_row = row_idx
            break
    if header_row is None:
        raise WorkbookProcessError(f"sheet「{sheet.title}」没有找到月清单表头。")

    data_start = header_row + 1
    total_row = None
    for row_idx in range(data_start, (sheet.max_row or data_start) + 1):
        value = clean_text(sheet.cell(row_idx, 1).value)
        if "合计" in value:
            total_row = row_idx
            break
    if total_row is None:
        raise WorkbookProcessError(f"sheet「{sheet.title}」没有找到合计行。")
    return header_row, data_start, total_row


def copy_row_format(sheet: Worksheet, source_row: int, target_row: int, max_col: int = 8) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def resize_data_area(sheet: Worksheet, data_start: int, total_row: int, required_rows: int) -> int:
    current_rows = total_row - data_start
    if required_rows <= 0:
        raise WorkbookProcessError("没有可写入的明细行。")

    if required_rows < current_rows:
        sheet.delete_rows(data_start + required_rows, current_rows - required_rows)
    elif required_rows > current_rows:
        insert_at = data_start + current_rows
        add_count = required_rows - current_rows
        sheet.insert_rows(insert_at, add_count)
        style_source = max(data_start, insert_at - 1)
        for row_idx in range(insert_at, insert_at + add_count):
            copy_row_format(sheet, style_source, row_idx)
    return data_start + required_rows


def unmerge_dynamic_ranges(sheet: Worksheet, data_start: int) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row >= data_start and merged_range.max_col <= 8:
            sheet.unmerge_cells(str(merged_range))


def normalize_data_row_styles(sheet: Worksheet, data_start: int, required_rows: int) -> None:
    for row_idx in range(data_start, data_start + required_rows):
        copy_row_format(sheet, data_start, row_idx)


def merge_range_if_needed(sheet: Worksheet, start_row: int, end_row: int, col: int) -> None:
    if end_row > start_row:
        sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)


def clear_data_rows(sheet: Worksheet, data_start: int, required_rows: int) -> None:
    for row_idx in range(data_start, data_start + required_rows):
        for col in range(1, 9):
            sheet.cell(row_idx, col).value = None


def write_month_sheet(sheet: Worksheet, school_name: str, meal_type: str, year: int, month: int, rows: list[OrderRow]) -> GeneratedSheet:
    if not rows:
        raise WorkbookProcessError("没有可写入的订单。")

    _header_row, data_start, total_row = find_table_layout(sheet)
    unmerge_dynamic_ranges(sheet, data_start)
    total_row = resize_data_area(sheet, data_start, total_row, len(rows))
    normalize_data_row_styles(sheet, data_start, len(rows))
    clear_data_rows(sheet, data_start, len(rows))

    rows_by_date: dict[date, list[OrderRow]] = defaultdict(list)
    for item in rows:
        rows_by_date[item.delivery_date].append(item)

    current_row = data_start
    month_total = Decimal("0")
    group_ranges: list[tuple[int, int]] = []
    for delivery_day in sorted(rows_by_date):
        day_rows = sorted(rows_by_date[delivery_day], key=lambda item: (item.order_no, item.row_no))
        day_total = sum((item.amount for item in day_rows), Decimal("0"))
        month_total += day_total
        group_start = current_row
        for idx, item in enumerate(day_rows):
            sheet.cell(current_row, 1).value = delivery_day.isoformat() if idx == 0 else None
            sheet.cell(current_row, 2).value = item.product
            sheet.cell(current_row, 3).value = item.description
            sheet.cell(current_row, 4).value = item.quantity
            sheet.cell(current_row, 5).value = item.unit
            sheet.cell(current_row, 6).value = item.price
            sheet.cell(current_row, 7).value = decimal_to_cell(item.amount)
            sheet.cell(current_row, 8).value = decimal_to_cell(day_total) if idx == 0 else None
            current_row += 1
        group_ranges.append((group_start, current_row - 1))

    for start_row, end_row in group_ranges:
        merge_range_if_needed(sheet, start_row, end_row, 1)
        merge_range_if_needed(sheet, start_row, end_row, 8)

    first_day = min(rows_by_date)
    last_day = max(rows_by_date)
    sheet.cell(1, 1).value = (
        f"光山县早晚餐食材配送供货月清单（{year}年{month}月份）"
        if meal_type == "evening"
        else f"光山县营养餐食材配送供货月清单（{year}年{month}月份）"
    )
    sheet.cell(2, 1).value = f"配送学校：{school_name}"
    sheet.cell(2, 5).value = f"日期：{zh_date(first_day)}-{zh_date(last_day)}"

    sheet.cell(total_row, 1).value = f"{month}月合计"
    for col in range(2, 7):
        sheet.cell(total_row, col).value = None
    sheet.cell(total_row, 7).value = decimal_to_cell(month_total)
    sheet.cell(total_row, 8).value = decimal_to_cell(month_total)
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)

    signature_row = total_row + 1
    if signature_row <= (sheet.max_row or signature_row):
        sheet.cell(signature_row, 1).value = "配送公司盖章："
        sheet.cell(signature_row, 4).value = "学校签字确认："
        sheet.cell(signature_row, 7).value = "学校盖章："

    sheet.print_area = f"A1:H{signature_row}"
    return GeneratedSheet(
        name=sheet.title,
        meal_type=meal_type,
        year=year,
        month=month,
        rows=len(rows),
        days=len(rows_by_date),
        total=month_total,
    )


def process_workbook(
    source_path: Path | list[Path],
    template_path: Path,
    output_path: Path,
    school_name: str | None = None,
    include_regular: bool = True,
    include_evening: bool = True,
    replace_existing: bool = False,
) -> tuple[str, list[GeneratedSheet], list[str]]:
    target_workbook = load_workbook(template_path)
    try:
        source_paths = source_path if isinstance(source_path, list) else [source_path]
        school = clean_text(school_name) or extract_school_name(target_workbook)
        if not school:
            raise WorkbookProcessError("学校名称为空。")

        orders = read_orders_from_sources(source_paths, school, include_regular, include_evening)
        if not orders:
            raise WorkbookProcessError(f"订单导出文件里没有匹配「{school}」或「{school}-早晚餐」的记录。")

        grouped: dict[tuple[str, int, int], list[OrderRow]] = defaultdict(list)
        for item in orders:
            grouped[(item.meal_type, item.delivery_date.year, item.delivery_date.month)].append(item)

        generated: list[GeneratedSheet] = []
        warnings: list[str] = []
        for (meal_type, year, month), rows in sorted(grouped.items(), key=lambda item: (item[0][1], item[0][2], item[0][0])):
            template_sheet = choose_template_sheet(target_workbook, meal_type, year, month)
            desired_name = base_sheet_name(meal_type, year, month)
            existing_sheet = find_sheet_by_title(target_workbook, desired_name)
            sheet_already_exists = existing_sheet is not None

            if sheet_already_exists and replace_existing:
                sheet = existing_sheet
            else:
                sheet = target_workbook.copy_worksheet(template_sheet)
                sheet.title = safe_sheet_name(desired_name, target_workbook, replace_existing=False)
                if sheet_already_exists:
                    warnings.append(f"已保留原有「{desired_name}」，新数据写入「{sheet.title}」。")

            generated.append(write_month_sheet(sheet, school, meal_type, year, month, rows))

        output_path.parent.mkdir(parents=True, exist_ok=True)
        target_workbook.save(output_path)
        return school, generated, warnings
    finally:
        target_workbook.close()


class AppHandler(BaseHTTPRequestHandler):
    server_version = "LedgerImportTool/1.0"

    def log_message(self, format: str, *args: Any) -> None:
        print(f"{self.address_string()} - {format % args}")

    def send_html(self, content: bytes, status: int = 200) -> None:
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def redirect_home(self) -> None:
        self.send_response(303)
        self.send_header("Location", "/")
        self.end_headers()

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/":
            self.send_html(render_form())
            return
        if parsed.path.startswith("/download/"):
            name = Path(unquote(parsed.path.removeprefix("/download/"))).name
            path = get_output_path(name, allowed_suffixes=(".xlsx", ".zip"))
            if not path.exists():
                self.send_error(404, "文件不存在")
                return
            content_type = mimetypes.guess_type(path.name)[0] or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            data = path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Disposition", f'attachment; filename="{quote(path.name)}"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        self.send_error(404)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path.startswith("/delete/"):
            try:
                name = unquote(parsed.path.removeprefix("/delete/"))
                path = get_saved_output_path(name)
                if path.exists():
                    path.unlink()
                self.redirect_home()
            except Exception as exc:
                self.send_html(render_error(str(exc)), status=400)
            return

        if parsed.path != "/process":
            self.send_error(404)
            return

        try:
            content_length = int(self.headers.get("Content-Length", "0"))
            fields, files = parse_multipart(self.headers, self.rfile.read(content_length))
            source_uploads = files.get("source_file", [])
            template_uploads = files.get("template_file", [])
            if not source_uploads or not template_uploads:
                raise WorkbookProcessError("请同时上传订单导出文件和目标模板文件。")

            UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
            job_id = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:8]
            job_dir = UPLOAD_DIR / job_id
            job_dir.mkdir(parents=True, exist_ok=True)

            source_paths: list[Path] = []
            for index, (source_name, source_bytes) in enumerate(source_uploads, start=1):
                if not source_name.lower().endswith(".xlsx"):
                    raise WorkbookProcessError("订单导出文件只支持 .xlsx。")
                source_path = job_dir / f"订单_{index}_{safe_filename(source_name)}"
                source_path.write_bytes(source_bytes)
                source_paths.append(source_path)

            template_paths: list[tuple[str, Path]] = []
            for index, (template_name, template_bytes) in enumerate(template_uploads, start=1):
                if not template_name.lower().endswith(".xlsx"):
                    raise WorkbookProcessError("目标模板文件只支持 .xlsx。")
                template_path = job_dir / f"模板_{index}_{safe_filename(template_name)}"
                template_path.write_bytes(template_bytes)
                template_paths.append((template_name, template_path))

            include_regular = "include_regular" in fields
            include_evening = "include_evening" in fields
            replace_existing = "replace_existing" in fields
            manual_school = fields.get("school_name") if len(template_paths) == 1 else None

            generated_workbooks: list[GeneratedWorkbook] = []
            errors: list[str] = []
            for index, (template_name, template_path) in enumerate(template_paths, start=1):
                output_name = f"配送月清单_导入结果_{safe_filename(Path(template_name).stem, f'学校{index}')}_{index}_{job_id}.xlsx"
                output_path = OUTPUT_DIR / output_name
                try:
                    school, sheets, warnings = process_workbook(
                        source_path=source_paths,
                        template_path=template_path,
                        output_path=output_path,
                        school_name=manual_school,
                        include_regular=include_regular,
                        include_evening=include_evening,
                        replace_existing=replace_existing,
                    )
                    generated_workbooks.append(
                        GeneratedWorkbook(
                            output_name=output_name,
                            school=school,
                            template_name=template_name,
                            sheets=sheets,
                            warnings=warnings,
                        )
                    )
                except Exception as exc:
                    errors.append(f"{template_name}：{exc}")

            if not generated_workbooks:
                raise WorkbookProcessError("没有生成任何目标清单。\n" + "\n".join(errors))

            zip_name = f"配送月清单_批量结果_{job_id}.zip"
            create_batch_zip([item.output_name for item in generated_workbooks], zip_name)
            self.send_html(render_batch_result(generated_workbooks, errors, zip_name))
        except Exception as exc:
            detail = str(exc)
            if not isinstance(exc, WorkbookProcessError):
                detail = f"{detail}\n\n{traceback.format_exc(limit=3)}"
            self.send_html(render_error(detail), status=400)


def run_server(host: str, port: int) -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    server = ThreadingHTTPServer((host, port), AppHandler)
    print(f"Listening on http://{host}:{port}")
    server.serve_forever()


def run_cli(args: argparse.Namespace) -> None:
    source = Path(args.source).expanduser().resolve()
    template = Path(args.template).expanduser().resolve()
    output = Path(args.output).expanduser().resolve() if args.output else OUTPUT_DIR / f"配送月清单_导入结果_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    school, sheets, warnings = process_workbook(
        source_path=source,
        template_path=template,
        output_path=output,
        school_name=args.school,
        include_regular=not args.no_regular,
        include_evening=not args.no_evening,
        replace_existing=args.replace_existing,
    )
    print(f"输出文件: {output}")
    print(f"配送学校: {school}")
    for warning in warnings:
        print(f"提示: {warning}")
    for sheet in sheets:
        meal = "早晚餐" if sheet.meal_type == "evening" else "普通营养餐"
        print(f"{sheet.name}: {meal}, {sheet.rows} 行, {sheet.days} 天, 合计 {format_decimal(sheet.total)}")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="配送月清单导入工具")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument("--source", help="订单导出 .xlsx，用于命令行处理")
    parser.add_argument("--template", help="目标月清单模板 .xlsx，用于命令行处理")
    parser.add_argument("--output", help="输出 .xlsx 路径")
    parser.add_argument("--school", help="学校名称，不填则从模板读取")
    parser.add_argument("--no-regular", action="store_true", help="不处理普通营养餐")
    parser.add_argument("--no-evening", action="store_true", help="不处理早晚餐")
    parser.add_argument("--replace-existing", action="store_true", help="同名月份 sheet 存在时覆盖")
    return parser


if __name__ == "__main__":
    parser = build_arg_parser()
    parsed_args = parser.parse_args()
    if parsed_args.source or parsed_args.template:
        if not parsed_args.source or not parsed_args.template:
            raise SystemExit("--source 和 --template 必须同时提供。")
        run_cli(parsed_args)
    else:
        run_server(parsed_args.host, parsed_args.port)
